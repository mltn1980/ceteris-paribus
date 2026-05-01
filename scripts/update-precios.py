import requests, json, re
from io import BytesIO
from openpyxl import load_workbook
from datetime import date
from bs4 import BeautifulSoup

CM_URL = "https://camaramercantil.com.uy/cereales-y-oleaginosas/"

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36'
}

def parse_price(val):
    if val is None:
        return None
    s = str(val).strip()
    m = re.match(r'^(\d+)/(\d+)$', s)
    if m:
        return int((int(m.group(1)) + int(m.group(2))) / 2)
    if re.match(r'^\d+$', s):
        return int(s)
    return None

# 1. Encontrar el ultimo xlsx
page = requests.get(CM_URL, headers=HEADERS, timeout=30)
print(f"Status: {page.status_code} — URL final: {page.url}")
soup = BeautifulSoup(page.content, 'html.parser')
all_links = [a['href'] for a in soup.find_all('a', href=True)]
print(f"Total links encontrados: {len(all_links)}")
xlsx_links = [h for h in all_links if h.lower().endswith('.xlsx')]
print(f"Links xlsx: {xlsx_links[:5]}")
links = sorted(set(xlsx_links))
if not links:
    raise RuntimeError("No se encontraron archivos xlsx en la pagina")

latest_url = links[-1]
print(f"Archivo: {latest_url}")

m = re.search(r'(\d{4})\.(\d{2})\.(\d{2})', latest_url)
fecha = f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else str(date.today())

# 2. Descargar y parsear Excel
resp = requests.get(latest_url, headers=HEADERS, timeout=60)
wb = load_workbook(BytesIO(resp.content), read_only=True, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))

# Ultima columna con datos (fila de fechas = row 2)
last_col = max((i for i, v in enumerate(rows[1]) if v is not None), default=0)

# 3. Extraer precios por cultivo
precios = {}
for row in rows:
    if not row[0]:
        continue
    desc = str(row[0]).strip()
    val  = parse_price(row[last_col])
    if val is None:
        continue

    if 'Industria_ZAFRA NUEVA' in desc:
        precios['soja']    = {'precio': val, 'desc': 'Industria zafra nueva - Mvd.'}
    elif 'PAN' in desc and 'ZAFRA NUEVA' in desc:
        precios['trigo']   = {'precio': val, 'desc': 'Zafra nueva - Mvd.'}
    elif 'Grado II' in desc and 'DISPONIBLE' in desc:
        precios['maiz']    = {'precio': val, 'desc': 'Grado II disponible - Mvd.'}
    elif 'bonificaci' in desc and 'ZAFRA NUEVA' in desc:
        precios['girasol'] = {'precio': val, 'desc': 'Industria zafra nueva - Mvd.'}
    elif 'Industria - Puesta' in desc and 'ZAFRA NUEVA' in desc:
        precios['colza']   = {'precio': val, 'desc': 'Industria zafra nueva - Mvd.'}

if not precios:
    raise RuntimeError("No se pudieron extraer precios del archivo")

# 4. Guardar JSON
output = {
    'fecha':  fecha,
    'fuente': 'Cámara Mercantil de Productos del País',
    'precios': precios
}
with open('data/precios.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"Precios al {fecha}:")
for k, v in precios.items():
    print(f"  {k}: USD {v['precio']}/tn")
