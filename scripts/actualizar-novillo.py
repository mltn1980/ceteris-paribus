#!/usr/bin/env python3
"""
actualizar-novillo.py
---------------------
Corre cada 2 días vía GitHub Actions.
Descarga el Excel de Novillo Tipo 2.0 de INAC y actualiza
data/novillo.json y data/novillo-tipo-serie.json.
No necesita Claude — los datos son numéricos directos.
"""

import io
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import requests
import xlrd

# ── Configuración ─────────────────────────────────────────────────────────────

NOVILLO_JSON = Path(__file__).parent.parent / "data" / "novillo.json"
SERIE_JSON   = Path(__file__).parent.parent / "data" / "novillo-tipo-serie.json"
TIMEOUT = 30
HEADERS  = {"User-Agent": "CeterisParibusBot/1.0 (+https://ceterisparibus.uy)"}
EXCEL_URL = "https://www.inac.uy/innovaportal/v/21699/10/innova.front/serie-mensual-novillo-tipo-20"

MESES_NOMBRE = {
    "ene": "enero", "feb": "febrero", "mar": "marzo",  "abr": "abril",
    "may": "mayo",  "jun": "junio",   "jul": "julio",  "ago": "agosto",
    "set": "setiembre", "sep": "setiembre",
    "oct": "octubre", "nov": "noviembre", "dic": "diciembre"
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def download_excel(url: str) -> bytes:
    r = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
    r.raise_for_status()
    return r.content


def _safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(round(float(val)))
    except (ValueError, TypeError):
        return None


def parse_excel(content: bytes) -> list:
    """Extrae la serie mensual desde el Excel de INAC (openpyxl, con fallback xlrd)."""
    import io as _io

    # Intentar openpyxl primero (xlsx)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
        ws = wb.active
        print(f"   → openpyxl: {ws.max_row} filas, {ws.max_column} columnas")

        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2:
                continue
            mes = row[0]
            nov = row[1] if len(row) > 1 else None
            vh  = row[2] if len(row) > 2 else None
            vai = row[3] if len(row) > 3 else None
            if not mes or _safe_int(nov) is None:
                continue
            mes_str = str(mes).strip()
            if not mes_str or mes_str in ("0", "0.0", "Mes"):
                continue
            rows.append({
                "mes":     mes_str,
                "novillo": _safe_int(nov),
                "vh":      _safe_int(vh) or 0,
                "vai":     _safe_int(vai) or 0,
            })

        if rows:
            print(f"   → {len(rows)} filas válidas, primera={rows[0]['mes']}, última={rows[-1]['mes']}")
            return rows
        print("   → openpyxl no encontró filas válidas, probando xlrd...")
    except Exception as e:
        print(f"   → openpyxl falló ({e}), probando xlrd...")

    # Fallback xlrd (xls legacy)
    MESES_CORTOS = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
                    7:"Jul",8:"Ago",9:"Set",10:"Oct",11:"Nov",12:"Dic"}
    wb2 = xlrd.open_workbook(file_contents=content)
    ws2 = wb2.sheet_by_index(0)
    print(f"   → xlrd: {ws2.nrows} filas, {ws2.ncols} columnas")
    rows = []
    for r in range(1, ws2.nrows):
        mes_cell = ws2.cell(r, 0)
        if mes_cell.ctype == xlrd.XL_CELL_DATE:
            dt = xlrd.xldate_as_datetime(mes_cell.value, wb2.datemode)
            mes = f"{MESES_CORTOS[dt.month]}-{str(dt.year)[2:]}"
        elif mes_cell.ctype in (xlrd.XL_CELL_TEXT, xlrd.XL_CELL_NUMBER):
            mes = str(mes_cell.value).strip()
            if not mes or mes in ("0", "0.0", "Mes"):
                continue
        else:
            continue
        nov = _safe_int(ws2.cell(r, 1).value if ws2.ncols > 1 else None)
        if nov is None:
            continue
        vh  = _safe_int(ws2.cell(r, 2).value if ws2.ncols > 2 else None) or 0
        vai = _safe_int(ws2.cell(r, 3).value if ws2.ncols > 3 else None) or 0
        rows.append({"mes": mes, "novillo": nov, "vh": vh, "vai": vai})

    return rows


def pct(new_val: int, old_val: int) -> float:
    if not old_val:
        return 0.0
    return round((new_val - old_val) / old_val * 100, 1)


def mes_label(mes_str: str) -> str:
    """Convierte 'Mar-26' → 'marzo 2026'."""
    partes = mes_str.split("-")
    if len(partes) != 2:
        return mes_str
    abrev = partes[0].lower()
    anio  = "20" + partes[1] if len(partes[1]) == 2 else partes[1]
    return MESES_NOMBRE.get(abrev, abrev) + " " + anio


def mismo_mes_año_anterior(serie: list, ultimo: dict) -> dict | None:
    """Encuentra la entrada del mismo mes del año anterior."""
    abrev_actual = ultimo["mes"].split("-")[0].lower()
    for d in reversed(serie[:-1]):
        if d["mes"].split("-")[0].lower() == abrev_actual:
            return d
    return None


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    forzar = os.environ.get("FORZAR", "false").lower() == "true"

    # 1. Leer estado actual
    current_serie = json.loads(SERIE_JSON.read_text(encoding="utf-8"))
    ultimo_mes_actual = current_serie[-1]["mes"] if current_serie else ""
    print(f"📂 Serie actual: último mes = {ultimo_mes_actual}")

    # 2. Descargar Excel
    print(f"📊 Descargando Excel de INAC...")
    try:
        content = download_excel(EXCEL_URL)
        print(f"   → {len(content):,} bytes descargados")
    except Exception as e:
        print(f"❌ Error al descargar: {e}")
        sys.exit(1)

    # 3. Identificar formato y parsear
    magic = content[:4]
    is_xlsx = magic == b'PK\x03\x04'
    is_xls  = magic[:2] == b'\xd0\xcf'
    fmt = 'xlsx' if is_xlsx else ('xls' if is_xls else f'desconocido ({magic.hex()})')
    print(f"   → Formato detectado: {fmt}", flush=True)
    if not is_xlsx and not is_xls:
        print(f"   → Primeros 200 bytes: {content[:200]}", flush=True)
        print("❌ El archivo no es Excel válido (¿responde HTML el servidor?)")
        sys.exit(1)

    try:
        serie = parse_excel(content)
        print(f"   → {len(serie)} filas parseadas", flush=True)
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"❌ Error al parsear Excel: {e}")
        sys.exit(1)

    if not serie:
        print("❌ Parse devolvió 0 filas.")
        sys.exit(1)
    print(f"   → Último mes: {serie[-1]['mes']}")

    # 4. Verificar si hay datos nuevos
    ultimo_mes_nuevo = serie[-1]["mes"]
    if ultimo_mes_nuevo == ultimo_mes_actual and not forzar:
        print(f"✅ Sin datos nuevos ({ultimo_mes_actual}). No se actualiza.")
        sys.exit(0)

    # 5. Calcular variaciones
    ultimo   = serie[-1]
    anterior = serie[-2] if len(serie) >= 2 else None
    año_ant  = mismo_mes_año_anterior(serie, ultimo)

    var_m_nov = pct(ultimo["novillo"], anterior["novillo"]) if anterior else 0.0
    var_a_nov = pct(ultimo["novillo"], año_ant["novillo"])  if año_ant  else None
    var_m_vh  = pct(ultimo["vh"],      anterior["vh"])      if anterior else 0.0
    var_a_vh  = pct(ultimo["vh"],      año_ant["vh"])       if año_ant  else None
    var_m_vai = pct(ultimo["vai"],     anterior["vai"])     if anterior else 0.0
    var_a_vai = pct(ultimo["vai"],     año_ant["vai"])      if año_ant  else None

    vh_pct  = round(ultimo["vh"]  / ultimo["novillo"] * 100)
    vai_pct = round(ultimo["vai"] / ultimo["novillo"] * 100)

    # 6. Construir novillo.json actualizado
    hoy = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    updated = {
        "valor":            ultimo["novillo"],
        "periodo":          mes_label(ultimo["mes"]),
        "actualizado":      hoy,
        "variacion":        var_m_nov,
        "vh":               ultimo["vh"],
        "vh_participacion": vh_pct,
        "vh_variacion":     var_m_vh,
        "vai":              ultimo["vai"],
        "vai_participacion": vai_pct,
        "vai_variacion":    var_m_vai
    }
    if var_a_nov is not None: updated["variacion_año"]    = var_a_nov
    if var_a_vh  is not None: updated["vh_variacion_año"] = var_a_vh
    if var_a_vai is not None: updated["vai_variacion_año"] = var_a_vai

    # 7. Escribir archivos
    NOVILLO_JSON.write_text(
        json.dumps(updated, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    SERIE_JSON.write_text(
        json.dumps(serie, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    print(f"📝 Actualizado:")
    print(f"   📅 {ultimo_mes_actual} → {ultimo_mes_nuevo}")
    print(f"   🐂 Novillo: {ultimo['novillo']} USD/cab ({'+' if var_m_nov >= 0 else ''}{var_m_nov}% vs mes ant)")
    if var_a_nov is not None:
        print(f"   📆 vs año ant: {'+' if var_a_nov >= 0 else ''}{var_a_nov}%")
    print(f"✅ Listo. Fecha: {hoy}")


if __name__ == "__main__":
    main()
